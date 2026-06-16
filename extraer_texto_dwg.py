#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
extraer_texto_dwg.py
--------------------
Escanea ARTES/BN, abre DWGs que empiezan por P, extrae todo el texto
(TEXT, MTEXT y atributos de bloques), compara contra vitros y mallas
de la BD Azure, y guarda los matches en mallas.rutas_arte.

Checkpoint JSON por carpeta: si AutoCAD se cae, retoma donde quedó.

Uso:
    py extraer_texto_dwg.py
    py extraer_texto_dwg.py --reiniciar
    py extraer_texto_dwg.py --solo-excel       (no abre AutoCAD, genera Excel del JSON)
"""

import os, sys, json, time, re, threading, argparse
from datetime import datetime

# ── dependencias ──────────────────────────────────────────────────────────────
try:
    import win32com.client, pythoncom
except ImportError:
    print("Falta pywin32.  pip install pywin32"); sys.exit(1)

try:
    import pyodbc
except ImportError:
    print("Falta pyodbc.  pip install pyodbc"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Falta openpyxl.  pip install openpyxl"); sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════
RUTA_BASE = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"

CONN_AZURE = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=agpcolombia.database.windows.net,1433;"
    "DATABASE=AGP_Ingenieria;"
    "UID=DevIngenieria;"
    "PWD=HiJE068i0LQVrwA;"
    "Encrypt=yes;"
    "TrustServerCertificate=no;"
    "Connection Timeout=30;"
)

CARPETAS_EXCLUIR = {"RHINO", "OBSOLETO", "GALSSJET", "_ARCHIVO"}
TIMEOUT_ARCHIVO  = 45       # segundos por archivo
GUARDAR_CADA_N   = 5        # checkpoint cada N archivos
CHECKPOINT_JSON  = "extraer_checkpoint.json"
RESULTADO_JSON   = "extraer_resultado.json"

# ══════════════════════════════════════════════════════════════════════════════
#  LOGGER
# ══════════════════════════════════════════════════════════════════════════════
class Logger:
    def _ts(self): return time.strftime("%H:%M:%S")
    def info(self, m):  print(f"{self._ts()}  {m}", flush=True)
    def ok(self, m):    print(f"{self._ts()}  [OK]  {m}", flush=True)
    def warn(self, m):  print(f"{self._ts()}  [!]   {m}", flush=True)
    def error(self, m): print(f"{self._ts()}  [ERR] {m}", flush=True)

log = Logger()

# ══════════════════════════════════════════════════════════════════════════════
#  BASE DE DATOS — carga de códigos y guardado de matches
# ══════════════════════════════════════════════════════════════════════════════
def db_conectar():
    for driver in ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"]:
        try:
            cs = CONN_AZURE.replace("ODBC Driver 17 for SQL Server", driver)
            return pyodbc.connect(cs, timeout=30)
        except Exception:
            continue
    log.error("No se pudo conectar a Azure SQL")
    return None

def db_cargar_codigos(conn):
    """
    Retorna tres sets con todos los códigos conocidos:
      vitros   -> set de strings (campo vitro en mallas.vitrojet)
      grandes  -> set de strings (campo codigo en mallas.grandes)
      pequenas -> set de strings (campo str(codigo) en mallas.pequenas)
    """
    cur = conn.cursor()

    cur.execute("SELECT vitro FROM mallas.vitrojet WHERE vitro IS NOT NULL")
    vitros = {r[0].strip() for r in cur.fetchall() if r[0]}

    cur.execute("SELECT codigo FROM mallas.grandes WHERE codigo IS NOT NULL")
    grandes = {r[0].strip() for r in cur.fetchall() if r[0]}

    cur.execute("SELECT CAST(codigo AS NVARCHAR) FROM mallas.pequenas WHERE descripcion IS NOT NULL OR cod_veh IS NOT NULL")
    pequenas = {r[0].strip() for r in cur.fetchall() if r[0]}

    cur.close()
    log.ok(f"Códigos cargados — vitros: {len(vitros):,}  grandes: {len(grandes):,}  pequeñas: {len(pequenas):,}")
    return vitros, grandes, pequenas

def db_guardar_matches(conn, matches):
    """
    matches = lista de dicts con claves:
      ruta_dwg, vehiculo, archivo, tipo_match, codigo
    Inserta ignorando duplicados (UNIQUE constraint).
    """
    if not matches:
        return 0
    cur = conn.cursor()
    sql = """
        IF NOT EXISTS (SELECT 1 FROM mallas.rutas_arte WHERE ruta_dwg=? AND tipo_match=? AND codigo=?)
        INSERT INTO mallas.rutas_arte (ruta_dwg, vehiculo, archivo, tipo_match, codigo)
        VALUES (?,?,?,?,?)
    """
    ok = 0
    for m in matches:
        try:
            cur.execute(sql,
                m["ruta_dwg"], m["tipo_match"], m["codigo"],
                m["ruta_dwg"], m["vehiculo"],   m["archivo"],
                m["tipo_match"], m["codigo"])
            ok += 1
        except Exception:
            pass
    try:
        conn.commit()
    except Exception:
        pass
    cur.close()
    return ok

def db_actualizar_rutas(conn, ruta_dwg, vitros_m, grandes_m, pequenas_m):
    """
    Actualiza ruta_dwg en la tabla correspondiente para cada match encontrado.
    Solo escribe si ruta_dwg está vacío (no sobreescribe ruta existente).
    """
    if not (vitros_m or grandes_m or pequenas_m):
        return
    cur = conn.cursor()
    try:
        for cod in vitros_m:
            cur.execute(
                "UPDATE mallas.vitrojet SET ruta_dwg=? WHERE vitro=? AND (ruta_dwg IS NULL OR ruta_dwg='')",
                ruta_dwg, cod)
        for cod in grandes_m:
            cur.execute(
                "UPDATE mallas.grandes SET ruta_dwg=? WHERE codigo=? AND (ruta_dwg IS NULL OR ruta_dwg='')",
                ruta_dwg, cod)
        for cod in pequenas_m:
            cur.execute(
                "UPDATE mallas.pequenas SET ruta_dwg=? WHERE CAST(codigo AS NVARCHAR)=? AND (ruta_dwg IS NULL OR ruta_dwg='')",
                ruta_dwg, cod)
        conn.commit()
    except Exception as e:
        log.warn(f"  ruta_dwg update: {e}")
        try: conn.rollback()
        except: pass
    finally:
        cur.close()

# ══════════════════════════════════════════════════════════════════════════════
#  CHECKPOINT / RESULTADO
# ══════════════════════════════════════════════════════════════════════════════
def cp_cargar(ruta):
    if not os.path.exists(ruta):
        return {}
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)
        procesados = data.get("procesados", {})
        log.ok(f"Checkpoint: {len(procesados)} archivo(s) ya procesados")
        return procesados
    except Exception as e:
        log.warn(f"Checkpoint corrupto ({e}), empezando de cero")
        return {}

def cp_guardar(ruta, procesados):
    try:
        tmp = ruta + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"ts": datetime.now().isoformat(), "procesados": procesados},
                      f, ensure_ascii=False, indent=2)
        os.replace(tmp, ruta)
    except Exception as e:
        log.warn(f"Error guardando checkpoint: {e}")
def resultado_guardar(ruta, datos):
    try:
        tmp = ruta + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        os.replace(tmp, ruta)
    except Exception as e:
        log.warn(f"Error guardando resultado: {e}")

# ══════════════════════════════════════════════════════════════════════════════
#  MATCHING — compara textos extraídos contra sets de códigos
# ══════════════════════════════════════════════════════════════════════════════
# Separadores típicos en textos de AutoCAD
_SEP = re.compile(r"[\s\r\n\t;,/\\|:\"'()\[\]{}<>]+")

def _tokens(texto):
    """Divide un texto en tokens limpios."""
    partes = _SEP.split(texto.strip())
    return {p.strip() for p in partes if len(p.strip()) >= 2}

def match_textos(textos, vitros, grandes, pequenas):
    """
    Dado el listado de textos de un DWG, retorna listas de matches:
      matches_vitro   = [codigo, ...]
      matches_grandes = [codigo, ...]
      matches_pequenas= [codigo, ...]
    """
    mv, mg, mp = set(), set(), set()

    for texto in textos:
        # Buscar match exacto con el texto completo (normalizado)
        t_norm = texto.strip()

        if t_norm in vitros:   mv.add(t_norm)
        if t_norm in grandes:  mg.add(t_norm)
        if t_norm in pequenas: mp.add(t_norm)

        # Buscar match por tokens (el texto puede tener basura alrededor)
        for tok in _tokens(texto):
            if tok in vitros:   mv.add(tok)
            if tok in grandes:  mg.add(tok)
            if tok in pequenas: mp.add(tok)

    return sorted(mv), sorted(mg), sorted(mp)

# ══════════════════════════════════════════════════════════════════════════════
#  MOTOR AUTOCAD
# ══════════════════════════════════════════════════════════════════════════════
class AutoCAD:
    def __init__(self):
        pythoncom.CoInitialize()
        try:
            self.app = win32com.client.GetActiveObject("AutoCAD.Application")
            self._suprimir()
            log.ok(f"AutoCAD conectado — versión {self.app.Version}")
        except Exception as e:
            log.error(f"No hay AutoCAD abierto: {e}")
            log.error("Abre AutoCAD (sin archivos) y vuelve a ejecutar.")
            sys.exit(1)

    def _suprimir(self):
        for v, val in [("XLOADCTL",0),("FILEDIA",0),("EXPERT",5),
                       ("PROXYSHOW",0),("BACKGROUNDPLOT",0)]:
            try: self.app.SetSystemVariable(v, val)
            except: pass

    def _restaurar(self):
        for v, val in [("XLOADCTL",2),("FILEDIA",1),("EXPERT",0),
                       ("PROXYSHOW",1),("BACKGROUNDPLOT",2)]:
            try: self.app.SetSystemVariable(v, val)
            except: pass

    def vivo(self):
        for _ in range(3):
            try:
                _ = self.app.Version
                return True
            except:
                time.sleep(1.0)
        return False

    def extraer_textos(self, ruta_abs, timeout=45):
        """
        Extrae TEXT, MTEXT y atributos de bloques del ModelSpace.
        Retorna (lista_textos, error_str).
        """
        result = [None]
        err    = [None]
        try:
            stream = pythoncom.CoMarshalInterThreadInterfaceInStream(
                pythoncom.IID_IDispatch, self.app)
        except Exception as e:
            return None, f"Marshal: {e}"

        def _worker():
            pythoncom.CoInitialize()
            doc = None
            try:
                app_h = win32com.client.Dispatch(
                    pythoncom.CoGetInterfaceAndReleaseStream(
                        stream, pythoncom.IID_IDispatch))
                doc = app_h.Documents.Open(ruta_abs, True)
                time.sleep(0.3)

                vistos = set()
                textos = []

                def _add(t):
                    t = t.strip()
                    if t and t not in vistos:
                        vistos.add(t)
                        textos.append(t)

                # ModelSpace + todos los layouts
                espacios = []
                try: espacios.append(doc.ModelSpace)
                except: pass
                try:
                    for i in range(doc.Layouts.Count):
                        try: espacios.append(doc.Layouts.Item(i).Block)
                        except: pass
                except: pass

                for esp in espacios:
                    try: n = esp.Count
                    except: continue
                    for i in range(n):
                        try:
                            ent = esp.Item(i)
                            nombre = ent.EntityName.upper()

                            if nombre in ("ACDBTEXT", "ACDBMTEXT"):
                                _add(ent.TextString)

                            elif nombre == "ACDBBLOCKREF":
                                # atributos del bloque
                                try:
                                    attrs = ent.GetAttributes()
                                    for a in attrs:
                                        try: _add(a.TextString)
                                        except: pass
                                except: pass

                        except: continue

                result[0] = textos
            except Exception as e:
                err[0] = str(e)[:120]
            finally:
                if doc is not None:
                    try: doc.Close(False)
                    except: pass
                try: pythoncom.CoUninitialize()
                except: pass

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        t.join(timeout)

        if t.is_alive():
            # cerrar documento colgado
            try:
                for i in range(self.app.Documents.Count - 1, -1, -1):
                    try:
                        d = self.app.Documents.Item(i)
                        if "Drawing1" not in d.Name:
                            d.Close(False); time.sleep(1.0); break
                    except: pass
            except: pass
            return None, "TIMEOUT"

        if err[0]:
            return None, err[0]
        return result[0], None

    def cerrar_docs(self):
        try:
            for i in range(self.app.Documents.Count - 1, -1, -1):
                try:
                    d = self.app.Documents.Item(i)
                    if "Drawing1" not in d.Name:
                        d.Close(False); time.sleep(0.1)
                except: pass
        except: pass

    def quit(self):
        self._restaurar()
        self.cerrar_docs()
        try: pythoncom.CoUninitialize()
        except: pass

# ══════════════════════════════════════════════════════════════════════════════
#  RECOLECTAR CANDIDATOS
# ══════════════════════════════════════════════════════════════════════════════
def archivo_ok(nombre):
    n = nombre.upper()
    return n.startswith("P") and n.endswith(".DWG") and not n.endswith(".BAK")

def recolectar(ruta_base):
    candidatos = []
    for dirpath, dirnames, filenames in os.walk(ruta_base):
        dirnames[:] = [d for d in dirnames
                       if d.upper() not in CARPETAS_EXCLUIR]

        if os.path.basename(dirpath).upper() == "ARTES":
            # Archivos sueltos en ARTES
            for f in filenames:
                if archivo_ok(f):
                    ruta_abs = os.path.join(dirpath, f)
                    vehiculo = os.path.relpath(dirpath, ruta_base).split(os.sep)[0]
                    candidatos.append({
                        "ruta_abs":    ruta_abs,
                        "ruta_rel":    os.path.relpath(ruta_abs, ruta_base),
                        "carpeta_rel": os.path.relpath(dirpath, ruta_base),
                        "vehiculo":    vehiculo,
                    })
            # Subcarpeta BN
            for sub in dirnames:
                if sub.strip().upper() == "BN":
                    ruta_bn = os.path.join(dirpath, sub)
                    try:
                        for f in os.listdir(ruta_bn):
                            if archivo_ok(f):
                                ruta_abs = os.path.join(ruta_bn, f)
                                vehiculo = os.path.relpath(dirpath, ruta_base).split(os.sep)[0]
                                candidatos.append({
                                    "ruta_abs":    ruta_abs,
                                    "ruta_rel":    os.path.relpath(ruta_abs, ruta_base),
                                    "carpeta_rel": os.path.relpath(ruta_bn, ruta_base),
                                    "vehiculo":    vehiculo,
                                })
                    except: pass

    return candidatos

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL DE REPORTE
# ══════════════════════════════════════════════════════════════════════════════
def guardar_excel(resultado):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MATCHES"

    HDR_COLOR  = "1F4E79"
    OK_COLOR   = "C6EFCE"
    NO_COLOR   = "FFF2CC"
    ERR_COLOR  = "FFCCCC"

    headers = ["VEHÍCULO","ARCHIVO","VITROS ENCONTRADOS",
               "MALLAS GRANDES","MALLAS PEQUEÑAS","TOTAL MATCHES",
               "ESTADO","TEXTOS EXTRAÍDOS","RUTA COMPLETA"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=HDR_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    anchos = [32,35,28,28,20,12,12,80,70]
    for col, w in enumerate(anchos, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)].width = w

    for fila, (_, info) in enumerate(sorted(resultado.items()), 2):
        estado   = info.get("estado","")
        vitros   = info.get("vitros",[])
        grandes  = info.get("grandes",[])
        pequenas = info.get("pequenas",[])
        total    = len(vitros)+len(grandes)+len(pequenas)

        if "ERROR" in estado or "TIMEOUT" in estado:
            color = ERR_COLOR
        elif total > 0:
            color = OK_COLOR
        else:
            color = NO_COLOR

        vals = [
            info.get("vehiculo",""),
            info.get("archivo",""),
            " | ".join(vitros),
            " | ".join(grandes),
            " | ".join(pequenas),
            total,
            estado,
            " | ".join(info.get("textos",[])),
            info.get("ruta_abs",""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(fila, col, v)
            c.fill      = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    nombre = f"Matches_DWG_{ts}.xlsx"
    wb.save(nombre)
    log.ok(f"Excel guardado: {nombre}")
    return nombre

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reiniciar",   action="store_true",
                        help="Borra checkpoint y empieza de cero")
    parser.add_argument("--solo-excel",  action="store_true",
                        help="No abre AutoCAD; genera Excel del JSON existente")
    parser.add_argument("--ruta",        type=str, default=None,
                        help="Ruta base alternativa para pruebas")
    parser.add_argument("--nombre",      type=str, default="red",
                        help="Prefijo para checkpoint/resultado (ej: local, red, test)")
    args = parser.parse_args()

    ruta_base      = args.ruta   if args.ruta   else RUTA_BASE
    nombre         = args.nombre if args.nombre else "red"
    checkpoint_f   = f"extraer_checkpoint_{nombre}.json"
    resultado_f    = f"extraer_resultado_{nombre}.json"

    log.info("=" * 65)
    log.info("EXTRACTOR DWG  →  VALIDADOR VITRO/MALLA  →  AZURE")
    log.info(f"Ruta base : {ruta_base}")
    log.info(f"Sesión    : {nombre}  (checkpoint: {checkpoint_f})")
    log.info("=" * 65)

    # ── modo solo-excel ────────────────────────────────────────────────────────
    if args.solo_excel:
        if not os.path.exists(resultado_f):
            log.error(f"No existe {resultado_f}"); return
        with open(resultado_f, "r", encoding="utf-8") as f:
            resultado = json.load(f)
        guardar_excel(resultado)
        return

    # ── reiniciar ──────────────────────────────────────────────────────────────
    if args.reiniciar:
        for fpath in [checkpoint_f, resultado_f]:
            if os.path.exists(fpath): os.remove(fpath)
        log.warn("Checkpoint eliminado. Empezando de cero.")

    # ── conectar BD y cargar códigos ───────────────────────────────────────────
    log.info("Conectando a Azure SQL...")
    conn = db_conectar()
    if conn is None:
        log.error("Sin conexión a BD. Saliendo."); sys.exit(1)
    vitros_set, grandes_set, pequenas_set = db_cargar_codigos(conn)

    # ── recolectar candidatos ──────────────────────────────────────────────────
    log.info("Buscando DWGs candidatos...")
    candidatos = recolectar(ruta_base)
    log.ok(f"Total candidatos: {len(candidatos)}")
    if not candidatos:
        log.warn("No se encontraron archivos. Revisa RUTA_BASE."); return

    # ── cargar checkpoint y resultado previo ───────────────────────────────────
    procesados = cp_cargar(checkpoint_f)
    resultado  = {}
    if os.path.exists(resultado_f):
        try:
            with open(resultado_f, "r", encoding="utf-8") as f:
                resultado = json.load(f)
        except: pass

    pendientes = [c for c in candidatos if c["ruta_rel"] not in procesados]

    log.info(f"Pendientes: {len(pendientes)} / {len(candidatos)}")
    if not pendientes:
        log.ok("Todo ya procesado. Generando Excel...")
        guardar_excel(resultado); return

    # ── conectar AutoCAD ──────────────────────────────────────────────────────
    motor = AutoCAD()

    t_inicio = datetime.now()
    n_ok = n_err = n_timeout = n_matches = 0

    try:
        for idx, c in enumerate(pendientes, 1):
            ruta_rel  = c["ruta_rel"]
            arch_name = os.path.basename(c["ruta_abs"])

            log.info(f"[{idx}/{len(pendientes)}] {ruta_rel}")

            if not motor.vivo():
                log.error("AutoCAD no responde. Guardando y saliendo.")
                break

            try:
                textos, error = motor.extraer_textos(c["ruta_abs"], TIMEOUT_ARCHIVO)
            except Exception as e:
                textos, error = None, str(e)[:120]

            if error == "TIMEOUT":
                estado = "TIMEOUT"; n_timeout += 1
                log.warn("  → TIMEOUT")
                time.sleep(2.0)
                vitros_m = grandes_m = pequenas_m = []
            elif error:
                estado = f"ERROR: {error}"; n_err += 1
                log.error(f"  → {estado}")
                time.sleep(1.5)
                vitros_m = grandes_m = pequenas_m = []
            else:
                estado = "OK"; n_ok += 1
                vitros_m, grandes_m, pequenas_m = match_textos(
                    textos, vitros_set, grandes_set, pequenas_set)
                total_m = len(vitros_m) + len(grandes_m) + len(pequenas_m)
                n_matches += total_m
                log.ok(f"  → {len(textos)} textos | "
                       f"vitros:{len(vitros_m)} grandes:{len(grandes_m)} "
                       f"peq:{len(pequenas_m)}")

                # 1) guardar matches en mallas.rutas_arte (historial completo)
                matches_bd = []
                for cod in vitros_m:
                    matches_bd.append({"ruta_dwg": c["ruta_abs"], "vehiculo": c["vehiculo"],
                                       "archivo": arch_name, "tipo_match": "VITRO",   "codigo": cod})
                for cod in grandes_m:
                    matches_bd.append({"ruta_dwg": c["ruta_abs"], "vehiculo": c["vehiculo"],
                                       "archivo": arch_name, "tipo_match": "GRANDE",  "codigo": cod})
                for cod in pequenas_m:
                    matches_bd.append({"ruta_dwg": c["ruta_abs"], "vehiculo": c["vehiculo"],
                                       "archivo": arch_name, "tipo_match": "PEQUENA", "codigo": cod})
                if matches_bd:
                    guardados = db_guardar_matches(conn, matches_bd)
                    log.info(f"     → {guardados} matches en rutas_arte")

                # 2) actualizar ruta_dwg en vitrojet / grandes / pequenas
                db_actualizar_rutas(conn, c["ruta_abs"], vitros_m, grandes_m, pequenas_m)

            # guardar en resultado JSON
            resultado[ruta_rel] = {
                "vehiculo":    c["vehiculo"],
                "carpeta_rel": c["carpeta_rel"],
                "archivo":     arch_name,
                "ruta_abs":    c["ruta_abs"],
                "estado":      estado,
                "textos":      textos or [],
                "vitros":      vitros_m   if estado == "OK" else [],
                "grandes":     grandes_m  if estado == "OK" else [],
                "pequenas":    pequenas_m if estado == "OK" else [],
            }
            procesados[ruta_rel] = estado

            if idx % GUARDAR_CADA_N == 0:
                cp_guardar(checkpoint_f, procesados)
                resultado_guardar(resultado_f, resultado)
                log.info(f"  [checkpoint {idx}/{len(pendientes)}]")

            motor.cerrar_docs()

    except KeyboardInterrupt:
        log.warn("Interrumpido por usuario.")
    finally:
        cp_guardar(checkpoint_f, procesados)
        resultado_guardar(resultado_f, resultado)
        motor.quit()
        try: conn.close()
        except: pass

    # ── resumen ────────────────────────────────────────────────────────────────
    dur = (datetime.now() - t_inicio).seconds
    log.info("=" * 65)
    log.ok(f"OK:{n_ok}  TIMEOUT:{n_timeout}  ERROR:{n_err}  MATCHES_TOTAL:{n_matches}")
    log.info(f"Tiempo: {dur//60}m {dur%60}s")
    log.info("=" * 65)

    guardar_excel(resultado)
    log.ok(f"JSON: {resultado_f}")

if __name__ == "__main__":
    main()
