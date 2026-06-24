# -*- coding: utf-8 -*-
"""
Sincronizador Excel → Azure SQL (mallas.*)
Estrategia: DELETE tabla + bulk INSERT con fast_executemany en lotes de 500.
Cada tabla abre y cierra su propia conexión — si Azure corta, reconecta solo.
El Excel es la fuente de verdad.
"""

import os, sys, time, datetime
import openpyxl

try:
    import pymssql as _pymssql
except ImportError:
    _pymssql = None

# Wrapper para compatibilidad con placeholders '?' (pyodbc) → '%s' (pymssql)
class _CursorWrap:
    def __init__(self, cur): self._c = cur
    def execute(self, sql, params=()):
        return self._c.execute(sql.replace("?", "%s"), params or ())
    def executemany(self, sql, seq):
        return self._c.executemany(sql.replace("?", "%s"), seq)
    def __getattr__(self, n): return getattr(self._c, n)

class _ConnWrap:
    def __init__(self, conn): self._c = conn
    def cursor(self): return _CursorWrap(self._c.cursor())
    def commit(self):   self._c.commit()
    def rollback(self): self._c.rollback()
    def close(self):    self._c.close()
    autocommit = False
    def __enter__(self): return self
    def __exit__(self, *a): self._c.__exit__(*a)

_EXCEL_DEFAULT = r"C:\Users\abotero\Downloads\LISTADO DE MALLAS Y GLASSJET 2025 (3).xlsx"
# Variable de entorno AGP_EXCEL para sobreescribir la ruta si es necesario.
EXCEL = os.environ.get("AGP_EXCEL", _EXCEL_DEFAULT)


MAX_VACIAS  = 500      # filas consecutivas col-A-vacía = fin real de datos
MAX_FILAS   = 2_000_000  # tope de seguridad extremo — nunca debería alcanzarse
CHUNK       = 150   # filas por lote en VALUES batch (límite 2100 params SQL Server)

# ─────────────────────────────────────────────────────────────────────────────
_log_fn = None

def log(msg, tag=""):
    if _log_fn:
        _log_fn(msg, tag)
    else:
        print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}", flush=True)

def clean(v, maxlen=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s[:maxlen] if maxlen and len(s) > maxlen else s

def conectar():
    if _pymssql is None:
        raise RuntimeError("pymssql no disponible — recompila el .exe")
    try:
        conn = _pymssql.connect(
            server="agpcolombia.database.windows.net",
            port=1433,
            user="DevIngenieria",
            password="HiJE068i0LQVrwA",
            database="AGP_Ingenieria",
            timeout=30,
            login_timeout=30,
            charset="UTF-8",
            tds_version="7.3",
        )
        return _ConnWrap(conn)
    except Exception as e:
        raise RuntimeError(f"No se pudo conectar a la base de datos.\nDetalle: {e}")

def _normalizar(s):
    """Normaliza para comparar nombres de hojas ignorando encoding de Ñ/tildes."""
    return (s.upper().strip()
             .replace('\xd1', 'N').replace('\xc3\xb1', 'N')
             .replace('Ñ', 'N').replace('?', 'N'))

def _hoja(wb, *candidatos):
    nombres_norm = {_normalizar(k): k for k in wb.sheetnames}
    for c in candidatos:
        if c in wb.sheetnames:
            return wb[c]
        norm = _normalizar(c)
        if norm in nombres_norm:
            return wb[nombres_norm[norm]]
    raise KeyError(f"No se encontró ninguna de: {candidatos}")

# ─────────────────────────────────────────────────────────────────────────────
#  Motor principal
# ─────────────────────────────────────────────────────────────────────────────

def _fila_blanca(row):
    """Fila completamente vacía (columna A sin valor). Única que cuenta como 'vacía' para cortar."""
    return not row or row[0] is None or str(row[0]).strip() == ""


def _bulk_sync(nombre, ws, sql_delete, sql_insert, build_row_fn, pk_index=0):
    """
    Lee la hoja, deduplica por pk_index, limpia la tabla y hace bulk INSERT
    en lotes de CHUNK. Reconecta automáticamente si Azure corta la conexión.
    pk_index=None → sin dedup (glassjet_viejo).
    """
    log(f"Leyendo {nombre} del Excel...")
    t0 = time.time()
    visto  = {}   # pk → tupla
    sin_pk = []
    vacias = 0
    n_leidas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        n_leidas += 1
        if n_leidas > MAX_FILAS:
            log(f"  WARN {nombre}: se alcanzó límite de {MAX_FILAS} filas — revisar Excel", "warn")
            break
        if _fila_blanca(row):
            vacias += 1
            if vacias >= MAX_VACIAS:
                break
            continue
        vacias = 0  # reset en cuanto hay algo en col A
        params = build_row_fn(row)
        if params is None:
            continue  # filtrada por calidad, pero NO cuenta como vacía
        if pk_index is None:
            sin_pk.append(params)
        else:
            visto[params[pk_index]] = params

    filas = list(visto.values()) if pk_index is not None else sin_pk
    dups  = sum(1 for _ in visto) - len(filas) if pk_index is not None else 0

    msg = f"  {nombre}: {len(filas)} filas leídas ({time.time()-t0:.1f}s)"
    if dups > 0:
        msg += f"  [{dups} duplicados en Excel ignorados]"
    log(msg)

    if not filas:
        log(f"  {nombre}: sin datos — omitido", "warn")
        return 0, 0

    # Extraer la parte VALUES (?,?,...) del sql_insert para construir multi-row
    # sql_insert tiene forma: "INSERT INTO tabla (c1,c2,...) VALUES (?,?,...,'literal')"
    # Necesitamos contar cuántos ? hay por fila
    n_params = sql_insert.count("?")

    def _multi_insert(cur, lote):
        """Un único INSERT con múltiples VALUE sets — rápido y sin bug de fast_executemany."""
        # Extraer la parte "INSERT INTO tabla (cols)" y la parte "VALUES (?,?,...)"
        val_start = sql_insert.upper().rfind("VALUES")
        prefix    = sql_insert[:val_start].rstrip()   # "INSERT INTO tabla (cols)"
        val_tpl   = sql_insert[val_start + 6:].strip()  # "(?,?,...,'literal')"

        # Para cada fila en el lote, la parte VALUES tiene que expandir los ?
        # Construimos: VALUES (row1),(row2),...
        rows_sql  = ",".join([val_tpl] * len(lote))
        sql_multi = f"{prefix} VALUES {rows_sql}"

        # Aplanar params: solo los ? (sin literales hardcodeados)
        params = []
        for f in lote:
            params.extend(f)
        cur.execute(sql_multi, params)

    n_lotes = -(-len(filas) // CHUNK)
    ok = err = 0
    cn = conectar()
    try:
        cn.cursor().execute(sql_delete)
        cn.commit()

        for i in range(0, len(filas), CHUNK):
            lote = filas[i:i + CHUNK]
            n    = i // CHUNK + 1
            cur  = cn.cursor()
            try:
                _multi_insert(cur, lote)
                cn.commit()
                ok += len(lote)
                log(f"  {nombre}: {n}/{n_lotes} — {ok:,} insertados", "dim")
            except Exception:
                log(f"  {nombre}: reconectando (lote {n})...", "warn")
                try: cn.close()
                except Exception: pass
                cn  = conectar()
                cur = cn.cursor()
                for f in lote:
                    try:
                        cur.execute(sql_insert, f)
                        ok += 1
                    except Exception as e:
                        err += 1
                        if err <= 3:
                            log(f"    error: {str(e)[:70]}", "warn")
                cn.commit()
            except Exception as e:
                try: cn.rollback()
                except Exception: pass
                log(f"  {nombre}: lote {n} fila a fila ({str(e)[:50]})...", "warn")
                cur2 = cn.cursor()
                for f in lote:
                    try:
                        cur2.execute(sql_insert, f)
                        ok += 1
                    except Exception as e2:
                        err += 1
                        if err <= 3:
                            log(f"    error: {str(e2)[:70]}", "warn")
                cn.commit()
    finally:
        try: cn.close()
        except Exception: pass

    log(f"  {nombre}: {ok:,} insertados  |  {err} errores",
        "ok" if err == 0 else "warn")
    return ok, err

# ─────────────────────────────────────────────────────────────────────────────
#  Motor MERGE (no borra asignados)
# ─────────────────────────────────────────────────────────────────────────────

def _bulk_sync_merge(nombre, ws, tabla, pk_col, columnas, build_row_fn, preservar_cond=None):
    """
    UPSERT rápido: bulk INSERT nuevos + bulk UPDATE existentes.
    EXCEPTO filas protegidas (estado=ASIGNADO del sistema de reservas).
    Nunca borra nada. Usa fast_executemany para mínimos round-trips a Azure.
    """
    if preservar_cond is None:
        preservar_cond = "estado = 'ASIGNADO'"

    log(f"Leyendo {nombre} del Excel (modo MERGE)...")
    t0 = time.time()
    visto = {}
    vacias = 0
    n_leidas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        n_leidas += 1
        if n_leidas > MAX_FILAS:
            log(f"  WARN {nombre}: se alcanzó límite de {MAX_FILAS} filas — revisar Excel", "warn")
            break
        if _fila_blanca(row):
            vacias += 1
            if vacias >= MAX_VACIAS:
                break
            continue
        vacias = 0  # reset en cuanto hay algo en col A
        params = build_row_fn(row)
        if params is None:
            continue  # filtrada por calidad, pero NO cuenta como vacía
        visto[params[0]] = params

    filas = list(visto.values())
    log(f"  {nombre}: {len(filas)} filas leídas ({time.time()-t0:.1f}s)")

    if not filas:
        log(f"  {nombre}: sin datos — omitido", "warn")
        return 0, 0

    cn = conectar()
    try:
        cur = cn.cursor()

        # 1. PKs protegidos (una sola consulta)
        cur.execute(f"SELECT {pk_col} FROM {tabla} WHERE {preservar_cond}")
        protegidos = {str(r[0]).strip() for r in cur.fetchall()}

        # 2. PKs existentes en BD (una sola consulta)
        cur.execute(f"SELECT {pk_col} FROM {tabla}")
        existentes = {str(r[0]).strip() for r in cur.fetchall()}

        # 3. Separar en nuevos (INSERT) vs existentes no protegidos (UPDATE)
        a_insertar = []
        a_actualizar = []
        saltados = 0
        for f in filas:
            pk_val = str(f[0]).strip()
            if pk_val in protegidos:
                saltados += 1
            elif pk_val in existentes:
                a_actualizar.append(f)
            else:
                a_insertar.append(f)

        cols_str = ", ".join(columnas)
        ph_str   = ", ".join(["?"] * len(columnas))
        upd_cols = [c for c in columnas if c != pk_col]
        upd_str  = ", ".join(f"{c}=?" for c in upd_cols)

        ok = err = 0

        # 4. Bulk INSERT (multi-row VALUES, igual que _bulk_sync — muy rápido)
        if a_insertar:
            val_tpl  = f"({ph_str})"
            n_lotes  = -(-len(a_insertar) // CHUNK)
            for i in range(0, len(a_insertar), CHUNK):
                lote = a_insertar[i:i + CHUNK]
                rows_sql  = ",".join([val_tpl] * len(lote))
                sql_multi = f"INSERT INTO {tabla} ({cols_str}) VALUES {rows_sql}"
                params    = []
                for f in lote:
                    params.extend(f)
                try:
                    cur.execute(sql_multi, params)
                    ok += len(lote)
                except Exception as e:
                    # fila a fila como fallback
                    for f in lote:
                        try:
                            cur.execute(f"INSERT INTO {tabla} ({cols_str}) VALUES ({ph_str})", list(f))
                            ok += 1
                        except Exception as e2:
                            err += 1
                            if err <= 3:
                                log(f"    insert error {f[0]}: {str(e2)[:60]}", "warn")

        # 5. Bulk UPDATE con fast_executemany (un lote por conexión, muy rápido)
        if a_actualizar:
            sql_upd = f"UPDATE {tabla} SET {upd_str} WHERE {pk_col}=?"
            cur.fast_executemany = True
            params_upd = [
                [f[i] for i, c in enumerate(columnas) if c != pk_col] + [f[0]]
                for f in a_actualizar
            ]
            try:
                cur.executemany(sql_upd, params_upd)
                ok += len(a_actualizar)
            except Exception as e:
                cur.fast_executemany = False
                log(f"  {nombre}: UPDATE bulk falló, fila a fila ({str(e)[:50]})...", "warn")
                for f in a_actualizar:
                    try:
                        upd_vals = [f[i] for i, c in enumerate(columnas) if c != pk_col]
                        cur.execute(sql_upd, upd_vals + [f[0]])
                        ok += 1
                    except Exception as e2:
                        err += 1
                        if err <= 3:
                            log(f"    update error {f[0]}: {str(e2)[:60]}", "warn")

        cn.commit()
        log(f"  {nombre}: {ok:,} ok ({len(a_insertar)} nuevos, {len(a_actualizar)} actualizados) "
            f"| {saltados} protegidos | {err} errores",
            "ok" if err == 0 else "warn")
        return ok, err
    finally:
        cn.close()


# ─────────────────────────────────────────────────────────────────────────────
#  Funciones por tabla
# ─────────────────────────────────────────────────────────────────────────────

def importar_vitrojet(ws):
    sql_i = (
        "INSERT INTO mallas.vitrojet "
        "(vitro,codigo_malla,tipo_malla,cod_completo,bnerig,vehiculo,version,ruta,cambio) "
        "VALUES (?,?,?,?,?,?,?,?,'excel')"
    )
    def build(r):
        vitro = clean(r[0], 30)
        if not vitro:
            return None
        malla    = clean(r[1], 30)  if len(r) > 1 else None
        vehiculo = clean(r[4], 200) if len(r) > 4 else None
        # Saltar filas con solo vitro y nada más (son reservas del sistema, no datos reales del Excel)
        if not any([malla, vehiculo,
                    clean(r[2], 100) if len(r) > 2 else None,
                    clean(r[5], 100) if len(r) > 5 else None]):
            return None
        tipo = "G" if malla and str(malla).upper().startswith("A-") else "P"
        return (
            vitro, malla, tipo,
            clean(r[2], 100)  if len(r) > 2 else None,  # cod_completo
            clean(r[3], 20)   if len(r) > 3 else None,  # bnerig
            vehiculo,
            clean(r[5], 100)  if len(r) > 5 else None,  # version
            clean(r[6], 500)  if len(r) > 6 else None,  # ruta
            'excel',
        )
    # Preservar asignaciones ya existentes: no borrar, hacer MERGE por vitro
    return _bulk_sync_merge("VITROJET", ws, "mallas.vitrojet", "vitro",
                            ["vitro","codigo_malla","tipo_malla","cod_completo",
                             "bnerig","vehiculo","version","ruta","cambio"],
                            build)


def importar_grandes(ws):
    def build(r):
        c = clean(r[0], 30) 
        if not c:
            return None
        desc = clean(r[2], 200) if len(r) > 2 else None
        cod  = clean(r[1], 30)  if len(r) > 1 else None
        if not any([desc, cod,
                    clean(r[3], 100) if len(r) > 3 else None,
                    clean(r[5], 100) if len(r) > 5 else None]):
            return None
        return (
            c,
            cod,
            desc,
            clean(r[3], 100)  if len(r) > 3 else None,
            clean(r[4], 20)   if len(r) > 4 else None,
            clean(r[5], 100)  if len(r) > 5 else None,
            clean(r[6], 300)  if len(r) > 6 else None,
            'excel',
        )
    return _bulk_sync_merge("MALLAS GRANDES", ws, "mallas.grandes", "codigo",
                            ["codigo","cod_veh","descripcion","pieza","tipo",
                             "version","concatenar","cambio"],
                            build)


def importar_pequenas(ws):
    def build(r):
        raw = clean(r[0])
        if not raw:
            return None
        try:
            codigo = int(float(raw))
        except Exception:
            return None
        desc = clean(r[2], 200) if len(r) > 2 else None
        if not desc:
            return None
        return (
            codigo,
            clean(r[1], 30)  if len(r) > 1 else None,
            desc,
            clean(r[3], 100) if len(r) > 3 else None,
            clean(r[4], 20)  if len(r) > 4 else None,
            clean(r[5], 100) if len(r) > 5 else None,
            clean(r[6], 300) if len(r) > 6 else None,
            clean(r[7], 100) if len(r) > 7 else None,
            'excel',
        )
    return _bulk_sync_merge("MALLAS PEQUEÑAS", ws, "mallas.pequenas", "codigo",
                            ["codigo","cod_veh","descripcion","pieza","tipo",
                             "version","concatenar","part_number","cambio"],
                            build)


def importar_pasta_plata(ws):
    def build(r):
        c = clean(r[0], 30)
        if not c:
            return None
        vehiculo = clean(r[2], 200)
        tipo     = clean(r[1], 20)
        if not any([vehiculo, tipo,
                    clean(str(r[3]), 30)  if r[3] is not None else None,
                    clean(str(r[4]), 100) if r[4] is not None else None]):
            return None
        return (
            c,
            tipo,
            vehiculo,
            clean(str(r[3]), 30)  if r[3] is not None else None,
            clean(str(r[4]), 100) if r[4] is not None else None,
            clean(str(r[5]), 100) if r[5] is not None else None,
            clean(r[6], 500)      if len(r) > 6 else None,
            clean(r[7], 200)      if len(r) > 7 else None,
            'excel',
        )
    return _bulk_sync_merge("PASTA DE PLATA", ws, "mallas.pasta_plata", "consecutivo",
                            ["consecutivo","tipo","vehiculo","cod_vehiculo",
                             "version","pieza","ruta_archivo","caso","cambio"],
                            build,
                            preservar_cond="1=0")  # pasta_plata no tiene sistema de reservas


def importar_vinilos(ws):
    def build(r):
        h = clean(r[0], 30)
        if not h:
            return None
        vehiculo = clean(r[1], 200) if len(r) > 1 else None
        if not any([vehiculo,
                    clean(str(r[2]), 30)  if len(r) > 2 and r[2] is not None else None,
                    clean(str(r[3]), 100) if len(r) > 3 and r[3] is not None else None]):
            return None
        tipo = clean(r[5], 20) if len(r) > 5 else None
        if tipo == "BN2":
            tipo = "BN"
        return (
            h,
            vehiculo,
            clean(str(r[2]), 30)  if len(r) > 2 and r[2] is not None else None,
            clean(str(r[3]), 100) if len(r) > 3 and r[3] is not None else None,
            clean(str(r[4]), 100) if len(r) > 4 and r[4] is not None else None,
            tipo,
            'excel',
        )
    return _bulk_sync_merge("VINILOS", ws, "mallas.vinilos", "herramental",
                            ["herramental","vehiculo","cod_vehiculo","version",
                             "pieza","tipo","cambio"],
                            build,
                            preservar_cond="1=0")  # vinilos no tiene sistema de reservas


def importar_glassjet_viejo(ws):
    sql_i = (
        "INSERT INTO mallas.glassjet_viejo "
        "(malla,glassjet,part_number,tipo,vehiculo,homologacion_vitro) "
        "VALUES (?,?,?,?,?,?)"
    )
    def build(r):
        if not any(v is not None for v in r[:6]):
            return None
        return (
            clean(str(r[0]), 50) if r[0] is not None else None,
            clean(str(r[1]), 50) if r[1] is not None else None,
            clean(r[2], 100),
            clean(r[3], 20),
            clean(r[4], 200),
            clean(r[5], 50),
        )
    return _bulk_sync("GLASSJET VIEJO", ws,
                      "DELETE FROM mallas.glassjet_viejo", sql_i, build,
                      pk_index=None)

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(log_fn=None):
    global _log_fn
    _log_fn = log_fn

    log("=== SINCRONIZADOR AGP Excel -> Azure SQL ===")

    if not os.path.isfile(EXCEL):
        log(f"ERROR: Excel no encontrado:\n  {EXCEL}", "err")
        return False

    log("Abriendo Excel...")
    t0 = time.time()
    try:
        wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    except Exception as e:
        log(f"ERROR al abrir Excel: {e}", "err")
        return False
    log(f"Excel cargado en {time.time()-t0:.1f}s")

    log("Verificando conexión Azure SQL...")
    try:
        _t = conectar(); _t.close()
    except Exception as e:
        log(str(e), "err")
        return False
    log("Conexión OK\n")

    TAREAS = [
        ("VITROJET",      importar_vitrojet,      ["VITROJET"]),
        ("GRANDES",       importar_grandes,        ["GRANDES"]),
        ("PEQUEÑAS",      importar_pequenas,       ["PEQUEÑAS", "PEQUE\xd1AS", "PEQUENAS", "PEQUE?AS"]),
        ("PASTA DE PLATA",importar_pasta_plata,    ["PASTA DE PLATA"]),
        ("GLASSJET VIEJO",importar_glassjet_viejo, ["GLASSJET VIEJO NO ALIMENTAR INV", "GLASSJET VIEJO"]),
        ("VINILOS",       importar_vinilos,        ["VINILOS"]),
    ]

    errores_total = 0
    try:
        for nombre_tarea, fn, candidatos in TAREAS:
            try:
                ws_hoja = _hoja(wb, *candidatos)
            except KeyError:
                log(f"  WARN: hoja '{nombre_tarea}' no encontrada en el Excel — omitida", "warn")
                continue
            try:
                _, e = fn(ws_hoja)
                errores_total += e
            except Exception as ex:
                log(f"  ERROR en {nombre_tarea}: {ex}", "err")
                errores_total += 1

        log(f"\n=== Completado en {time.time()-t0:.0f}s ===",
            "ok" if errores_total == 0 else "warn")

        # Resincronizar secuencias para que el consecutivo siga desde el MAX real
        try:
            from asignaciones import sincronizar_secuencias
            sincronizar_secuencias()
            log("  Secuencias de consecutivo actualizadas", "ok")
        except Exception as _se:
            log(f"  WARN secuencias: {_se}", "warn")

        cn = conectar()
        cur = cn.cursor()
        for tabla in ["mallas.vitrojet", "mallas.grandes", "mallas.pequenas",
                      "mallas.pasta_plata", "mallas.glassjet_viejo", "mallas.vinilos"]:
            cur.execute(f"SELECT COUNT(*) FROM {tabla}")
            log(f"  {tabla}: {cur.fetchone()[0]:,} registros", "dim")
        cn.close()

    except Exception as e:
        log(f"ERROR FATAL: {e}", "err")
        import traceback; traceback.print_exc()
        return False
    finally:
        wb.close()

    return errores_total == 0


if __name__ == "__main__":
    sys.exit(0 if main() else 1)
