#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
BÚSQUEDA DE TEXTO EN DWGs — AGP
Ruta  : \\192.168.2.37\Ingenieria\PRODUCCION\AGP PLANOS TECNICOS\CHEVROLET\CHEVROLET SUBURBAN 4D U 07--239
Busca : CS7_-08-GS-XXXX20-C  (y variantes: sin case, sin _ ni -)
Reporta: cuáles archivos tienen el texto y cuáles no, con la carpeta donde están.
"""

import os
import sys
import time
import threading
from datetime import datetime

# ──────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────
RUTA_BASE     = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS\JEEP\JEEP GRAND CHEROKEE 4D U 2014 -- 342"
ARCHIVO_EXCEL = f"Busqueda_Texto_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
TIMEOUT_ARCHIVO = 40   # segundos máximo por archivo (subido porque algunos archivos son pesados)

# ¿Qué buscamos?
# Los cajetines de offset/BN en los DWGs de AGP tienen estas etiquetas:
#   OFFSET  (el valor del offset)
#   BN+D    (BN más tolerancia)
#   BN INT  (BN interior)
#   ACERO   (valor de acero)
#
# Un archivo SE CONSIDERA CON CAJETÍN si tiene AL MENOS estas 3 etiquetas.
# No importa el valor numérico, solo que existan las etiquetas.
# Cambia los patrones aquí si en tu carpeta el formato es diferente.
# Para ser COMPLETO el archivo debe tener TODO esto:
#
#  ① Etiqueta de offset  → "OFFSET" o "OFF:"
#  ② Etiqueta de BN      → "BN+D" o "BN INT" o "BN:"
#  ③ Cajetín 1           → valor 67.0
#  ④ Cajetín 2           → valor 230.0
#  ⑤ Cajetín 3           → valor 64.0
#  ⑥ Cajetín 4           → valor 45.0
#  ⑦ Cajetín 5           → valor 160.0
#
# Cada item se busca de forma independiente en cualquier entidad del archivo.
# Si falta cualquiera → PARCIAL (con detalle de qué falta).
CAJETINES = [
    {"nombre": "Etiqueta OFFSET",        "variantes": ["OFFSET", "OFF:"]},
    {"nombre": "Etiqueta BN",            "variantes": ["BN+D", "BN INT", "BN:"]},
    {"nombre": "Cajetín 1  BN: 67.0",   "variantes": ["67.0"]},
    {"nombre": "Cajetín 2  BN: 230.0",  "variantes": ["230.0"]},
    {"nombre": "Cajetín 3  BN: 64.0",   "variantes": ["64.0"]},
    {"nombre": "Cajetín 4  BN: 45.0",   "variantes": ["45.0"]},
    {"nombre": "Cajetín 5  BN: 160.0",  "variantes": ["160.0"]},
]
# ──────────────────────────────────────────────────────────

try:
    import win32com.client
    import pythoncom
except ImportError:
    print("Falta pywin32. Ejecuta: pip install pywin32")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────
# LOGGER
# ──────────────────────────────────────────────────────────
class Logger:
    def _ts(self): return time.strftime("%H:%M:%S")
    def info(self, m):  print(f"{self._ts()}  {m}")
    def warn(self, m):  print(f"{self._ts()}  [!] {m}")
    def error(self, m): print(f"{self._ts()}  [X] {m}")
    def ok(self, m):    print(f"{self._ts()}  [✓] {m}")

log = Logger()


# ──────────────────────────────────────────────────────────
# LÓGICA DE COINCIDENCIA POR CAJETÍN
#
# Para cada texto de entidad, revisamos si contiene alguno de
# los patrones pendientes. Usamos normalización:
#   - sin espacios extra
#   - sin importar mayúsculas
#   - "BN: 67.0+5" contiene "BN:67.0" → cumple
# ──────────────────────────────────────────────────────────
def _norm(t):
    """Mayúsculas y sin espacios extra — para comparar sin importar case ni espacios."""
    return t.upper().strip()

# Pre-calcular las variantes normalizadas una sola vez
_VARIANTES_NORM = [
    [_norm(v) for v in c["variantes"]]
    for c in CAJETINES
]

def _cajetines_en_texto(texto, pendientes_idx):
    """
    Dado un texto de entidad y un set de índices pendientes,
    retorna qué índices de cajetines están presentes en ese texto.
    Un cajetín se cumple si el texto contiene CUALQUIERA de sus variantes.
    """
    if not texto:
        return set()
    t_norm = _norm(texto)
    encontrados = set()
    for idx in pendientes_idx:
        for variante in _VARIANTES_NORM[idx]:
            if variante in t_norm:
                encontrados.add(idx)
                break   # con una variante que cumpla es suficiente
    return encontrados


# ──────────────────────────────────────────────────────────
# MOTOR AUTOCAD
# ──────────────────────────────────────────────────────────
class AutoCAD:
    def __init__(self):
        pythoncom.CoInitialize()
        try:
            self.app = win32com.client.GetActiveObject("AutoCAD.Application")
            self._suprimir_dialogs()
            log.ok(f"AutoCAD conectado: {self.app.Version}")
        except Exception as e:
            log.error(f"No hay AutoCAD abierto: {e}")
            log.error("Abre AutoCAD (sin archivos) y vuelve a ejecutar.")
            sys.exit(1)

    def _suprimir_dialogs(self):
        for var, val in [("XLOADCTL", 0), ("FILEDIA", 0), ("EXPERT", 5), ("PROXYSHOW", 0)]:
            try:
                self.app.SetSystemVariable(var, val)
            except Exception:
                pass
            
    def _restaurar_dialogs(self):
        for var, val in [("XLOADCTL", 2), ("FILEDIA", 1), ("EXPERT", 0), ("PROXYSHOW", 1)]:
            try:
                self.app.SetSystemVariable(var, val)
            except Exception:
                pass

    def vivo(self):
        for _ in range(3):
            try:
                _ = self.app.Version
                return True
            except Exception:
                time.sleep(1.0)
        return False

    def buscar_texto_con_timeout(self, ruta, timeout=TIMEOUT_ARCHIVO):
        """
        Abre el DWG, busca texto en todas las entidades y cierra.
        Todo en hilo secundario con timeout.
        Retorna (lista_de_coincidencias, error_str)
        """
        ruta_abs = os.path.abspath(ruta)
        result   = [None]   # lista de textos coincidentes, o None si error
        err      = [None]
        try:
            stream = pythoncom.CoMarshalInterThreadInterfaceInStream(
                pythoncom.IID_IDispatch, self.app
            )
        except Exception as e:
            return None, f"Marshal falló: {e}"

        def _buscar_en_coleccion(coleccion, pendientes):
            """
            Itera entidades buscando los cajetines pendientes.
            Early exit: para en cuanto encuentra TODOS los pendientes.
            Retorna set de índices encontrados en esta colección.
            """
            hallados = set()
            try:
                count = coleccion.Count
            except Exception:
                return hallados
            aun_pendientes = set(pendientes)
            for i in range(count):
                if not aun_pendientes:
                    break   # ya encontró todos, no seguir
                try:
                    ent  = coleccion.Item(i)
                    tipo = ent.EntityName
                    if tipo in ("AcDbText", "AcDbMText", "AcDbAttributeDefinition"):
                        encontrados = _cajetines_en_texto(ent.TextString, aun_pendientes)
                        hallados   |= encontrados
                        aun_pendientes -= encontrados
                    elif tipo == "AcDbBlockReference":
                        try:
                            for attr in ent.GetAttributes():
                                encontrados = _cajetines_en_texto(attr.TextString, aun_pendientes)
                                hallados   |= encontrados
                                aun_pendientes -= encontrados
                                if not aun_pendientes:
                                    break
                        except Exception:
                            pass
                except Exception:
                    continue
            return hallados

        def _worker():
            pythoncom.CoInitialize()
            doc = None
            try:
                app_hilo = win32com.client.Dispatch(
                    pythoncom.CoGetInterfaceAndReleaseStream(
                        stream, pythoncom.IID_IDispatch
                    )
                )
                doc = app_hilo.Documents.Open(ruta_abs, True)  # read-only
                time.sleep(0.1)

                todos_idx   = set(range(len(CAJETINES)))
                hallados    = set()

                # Buscar en ModelSpace
                try:
                    hallados |= _buscar_en_coleccion(doc.ModelSpace, todos_idx - hallados)
                except Exception:
                    pass

                # Si faltan cajetines, buscar en layouts (PaperSpace)
                if hallados != todos_idx:
                    try:
                        layouts = doc.Layouts
                        for li in range(layouts.Count):
                            if hallados == todos_idx:
                                break
                            try:
                                layout = layouts.Item(li)
                                if layout.Name.upper() == "MODEL":
                                    continue
                                hallados |= _buscar_en_coleccion(
                                    layout.Block, todos_idx - hallados
                                )
                            except Exception:
                                continue
                    except Exception:
                        pass

                # Resultado: dict con qué cajetines se encontraron
                result[0] = {
                    "hallados":  sorted(hallados),
                    "completo":  hallados == todos_idx,
                    "detalle":   [CAJETINES[i]["nombre"] + ": " + "/".join(CAJETINES[i]["variantes"])
                                  for i in sorted(hallados)],
                    "faltantes": [CAJETINES[i]["nombre"] + " (" + "/".join(CAJETINES[i]["variantes"]) + ")"
                                  for i in sorted(todos_idx - hallados)],
                }

            except Exception as e:
                err[0] = str(e)[:100]
            finally:
                if doc is not None:
                    try:
                        doc.Close(False)
                        # sin sleep: archivo de solo lectura, no hay datos que flushar
                    except Exception:
                        pass
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        t = threading.Thread(target=_worker, daemon=True)
        t.start()
        t.join(timeout)

        if t.is_alive():
            log.warn(f"  TIMEOUT ({timeout}s) — saltando archivo colgado")
            try:
                for i in range(self.app.Documents.Count - 1, -1, -1):
                    try:
                        doc = self.app.Documents.Item(i)
                        if "Drawing1" not in doc.Name:
                            doc.Close(False)
                            time.sleep(1.0)
                            break
                    except Exception:
                        pass
            except Exception:
                pass
            return None, "TIMEOUT — archivo colgado"

        if err[0]:
            return None, err[0]

        return result[0], None

    def cerrar_docs_abiertos(self):
        try:
            if self.app.Documents.Count <= 1:
                return
            for i in range(self.app.Documents.Count - 1, -1, -1):
                try:
                    doc = self.app.Documents.Item(i)
                    if "Drawing1" not in doc.Name:
                        doc.Close(False)
                        time.sleep(0.05)
                except Exception:
                    pass
        except Exception:
            pass

    def quit(self):
        self._restaurar_dialogs()
        self.cerrar_docs_abiertos()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ──────────────────────────────────────────────────────────
# RECOLECTAR TODOS LOS DWGs RECURSIVAMENTE
# ──────────────────────────────────────────────────────────
CARPETAS_EXCLUIR = ["RHINO", "OBSOLETO", "GALSSJET"]   # excluir: no son DWGs de AutoCAD válidos

def recolectar_dwgs(ruta_base):
    """
    Recorre recursivamente ruta_base y retorna lista de (ruta_relativa_carpeta, ruta_abs).
    Excluye carpetas cuyo nombre contenga alguna de las palabras en CARPETAS_EXCLUIR.
    """
    dwgs = []
    for dirpath, dirnames, filenames in os.walk(ruta_base):
        carpeta_rel = os.path.relpath(dirpath, ruta_base)

        # Excluir carpetas no deseadas (en cualquier nivel del path)
        partes = carpeta_rel.upper().replace("\\", "/").split("/")
        if any(excl in parte for excl in CARPETAS_EXCLUIR for parte in partes):
            dirnames.clear()   # evita que os.walk baje más dentro de esta carpeta
            continue

        if carpeta_rel == ".":
            carpeta_rel = "(raíz)"

        for fname in filenames:
            if fname.lower().endswith(".dwg"):
                ruta_abs = os.path.join(dirpath, fname)
                dwgs.append((carpeta_rel, fname, ruta_abs))

    return sorted(dwgs, key=lambda x: (x[0], x[1]))


# ──────────────────────────────────────────────────────────
# ESCANEO PRINCIPAL
# ──────────────────────────────────────────────────────────
def escanear(ruta_base, motor):
    dwgs = recolectar_dwgs(ruta_base)
    if not dwgs:
        log.warn(f"No se encontraron DWGs en: {ruta_base}")
        return []

    log.info(f"DWGs encontrados: {len(dwgs)}")
    log.info(f"Cajetines requeridos ({len(CAJETINES)}): " +
             "  |  ".join(c["nombre"] + ": " + "/".join(c["variantes"]) for c in CAJETINES))
    log.info("-" * 70)

    resultados = []
    for idx, (carpeta, nombre, ruta_abs) in enumerate(dwgs, 1):
        log.info(f"[{idx}/{len(dwgs)}] {carpeta}  /  {nombre}")

        try:
            res, error = motor.buscar_texto_con_timeout(ruta_abs)
            motor.cerrar_docs_abiertos()
        except Exception as e:
            log.warn(f"  → ERROR inesperado: {str(e)[:60]} — continuando...")
            motor.cerrar_docs_abiertos()
            res, error = None, str(e)[:80]

        if res is None:
            if not motor.vivo():
                log.error("AutoCAD dejó de responder.")
                return resultados
            estado    = "ERROR"
            detalle   = error or "No se pudo abrir"
            faltantes = ""
            encontro  = False
            log.warn(f"  → ERROR: {detalle[:60]}")
            time.sleep(1.5)
        elif res["completo"]:
            estado    = "COMPLETO"
            detalle   = f"Todos los cajetines presentes ({len(CAJETINES)}/{len(CAJETINES)})"
            faltantes = ""
            encontro  = True
            log.ok(f"  → COMPLETO — todos los {len(CAJETINES)} cajetines encontrados")
        elif res["hallados"]:
            estado    = "PARCIAL"
            detalle   = f"{len(res['hallados'])}/{len(CAJETINES)} cajetines: " + " | ".join(res["detalle"])
            faltantes = " | ".join(res["faltantes"])
            encontro  = False
            log.warn(f"  → PARCIAL: {len(res['hallados'])}/{len(CAJETINES)} — faltan: {faltantes[:60]}")
        else:
            estado    = "NO ENCONTRADO"
            detalle   = "Ningún cajetín presente"
            faltantes = " | ".join(res["faltantes"])
            encontro  = False
            log.info(f"  → ningún cajetín")

        resultados.append({
            "carpeta":   carpeta,
            "archivo":   nombre,
            "ruta":      ruta_abs,
            "estado":    estado,
            "encontro":  encontro,
            "detalle":   detalle,
            "faltantes": faltantes,
        })

    return resultados


# ──────────────────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────────────────
C = {
    "hdr":        "1F3864",
    "titulo":     "2E75B6",
    "white":      "FFFFFF",
    "alt":        "EEF4FF",
    "encontrado": "C6EFCE",
    "no_enc":     "FFCCCC",
    "error":      "FFC7CE",
}
_thin   = Side(style="thin", color="BBBBBB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _ch(cell, bg=None):
    cell.font      = Font(name="Arial", bold=True, color=C["white"], size=10)
    cell.fill      = PatternFill("solid", start_color=bg or C["hdr"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border

def _cd(cell, alt=False, center=False, bold=False, bg=None, mono=False, color_txt=None):
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          size=8 if mono else 9, bold=bold,
                          color=color_txt or "000000")
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = _border
    fill = bg or (C["alt"] if alt else None)
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)


def crear_excel(resultados, ruta_salida):
    if not resultados:
        log.warn("Sin datos para generar Excel.")
        return

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    completos    = [r for r in resultados if r["estado"] == "COMPLETO"]
    parciales    = [r for r in resultados if r["estado"] == "PARCIAL"]
    no_enc       = [r for r in resultados if r["estado"] == "NO ENCONTRADO"]
    errores      = [r for r in resultados if r["estado"] == "ERROR"]

    C["parcial"]  = "FFEB9C"   # amarillo para parcial
    C["completo"] = "C6EFCE"   # verde para completo

    HEADERS = ["#", "CARPETA", "ARCHIVO", "ESTADO",
               "CAJETINES PRESENTES", "CAJETINES FALTANTES", "RUTA COMPLETA"]
    ANCHOS  = [5, 45, 38, 16, 65, 55, 85]

    patrones_str = "  |  ".join(f"{c['nombre']}: {'/'.join(c['variantes'])}" for c in CAJETINES)

    def _escribir_hoja(ws, filas, titulo_texto, bg_titulo=None):
        n = len(HEADERS)
        ws.sheet_view.showGridLines = False

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c = ws.cell(1, 1, titulo_texto)
        c.font      = Font(name="Arial", size=13, bold=True, color=C["white"])
        c.fill      = PatternFill("solid", start_color=bg_titulo or C["hdr"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        m = ws.cell(2, 1,
                    f"Cajetines requeridos: {patrones_str}   |   "
                    f"Total en esta hoja: {len(filas)}   |   {fecha}")
        m.font      = Font(name="Arial", size=9, italic=True, color="555555")
        m.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16

        for col, h in enumerate(HEADERS, 1):
            _ch(ws.cell(3, col, h), bg=bg_titulo or C["titulo"])
        ws.row_dimensions[3].height = 22

        for i, r in enumerate(filas, 1):
            row = i + 3
            alt = i % 2 == 0
            est = r["estado"]
            bg_est = {"COMPLETO":      C["completo"],
                      "PARCIAL":       C["parcial"],
                      "NO ENCONTRADO": C["no_enc"],
                      "ERROR":         C["error"]}.get(est, C["alt"])

            _cd(ws.cell(row, 1, i),              center=True, alt=alt)
            _cd(ws.cell(row, 2, r["carpeta"]),   alt=alt)
            _cd(ws.cell(row, 3, r["archivo"]),   alt=alt)
            _cd(ws.cell(row, 4, est),            center=True, bold=True, bg=bg_est)
            _cd(ws.cell(row, 5, r["detalle"]),   alt=alt)
            _cd(ws.cell(row, 6, r.get("faltantes", "")),
                alt=alt, bg=C["error"] if r.get("faltantes") else None)
            _cd(ws.cell(row, 7, r["ruta"]),      alt=alt, mono=True, color_txt="0070C0")

        for i, w in enumerate(ANCHOS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A4"

    # Hoja RESUMEN
    ws_res = wb.create_sheet("RESUMEN")
    ws_res.sheet_view.showGridLines = False
    ws_res.merge_cells("A1:D1")
    c = ws_res.cell(1, 1, "BÚSQUEDA DE TEXTO EN DWGs — RESUMEN")
    c.font = Font(name="Arial", size=14, bold=True, color=C["white"])
    c.fill = PatternFill("solid", start_color=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 32

    ws_res.merge_cells("A2:D2")
    m = ws_res.cell(2, 1,
                    f"Cajetines: {patrones_str}   |   "
                    f"Ruta: {RUTA_BASE}   |   {fecha}")
    m.font = Font(name="Arial", size=9, italic=True, color="555555")
    m.alignment = Alignment(horizontal="left", vertical="center")
    ws_res.row_dimensions[2].height = 16

    for col, h in enumerate(["CATEGORÍA", "CANTIDAD", "%", "DESCRIPCIÓN"], 1):
        _ch(ws_res.cell(3, col, h))
    ws_res.row_dimensions[3].height = 22

    total = len(resultados)
    resumen_filas = [
        ("COMPLETO",      len(completos), C["completo"], f"Todos los {len(CAJETINES)} cajetines presentes"),
        ("PARCIAL",       len(parciales), C["parcial"],  "Algunos cajetines presentes, no todos"),
        ("NO ENCONTRADO", len(no_enc),    C["no_enc"],   "Ningún cajetín presente"),
        ("ERROR",         len(errores),   C["error"],    "Archivos que no se pudieron abrir"),
        ("TOTAL",         total,          None,          "Total de archivos DWG escaneados"),
    ]
    for i, (cat, cnt, bg, desc) in enumerate(resumen_filas, 4):
        alt = i % 2 == 0
        _cd(ws_res.cell(i, 1, cat),  bold=True, bg=bg, center=True)
        _cd(ws_res.cell(i, 2, cnt),  bold=True, center=True, bg=bg)
        pct = f"{cnt/total*100:.1f}%" if total else "0%"
        _cd(ws_res.cell(i, 3, pct),  center=True, alt=alt)
        _cd(ws_res.cell(i, 4, desc), alt=alt)

    for col, w in zip([1, 2, 3, 4], [20, 12, 10, 55]):
        ws_res.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws_res.freeze_panes = "A4"

    # Hoja TODOS
    _escribir_hoja(wb.create_sheet("TODOS LOS ARCHIVOS"), resultados,
                   f"TODOS LOS ARCHIVOS ({len(resultados)} DWGs)")

    # Hoja COMPLETOS (tienen todos los cajetines)
    if completos:
        _escribir_hoja(wb.create_sheet(f"COMPLETOS ({len(completos)})"), completos,
                       f"COMPLETOS — TODOS LOS CAJETINES PRESENTES ({len(completos)})",
                       bg_titulo="375623")

    # Hoja PARCIALES (tienen algunos cajetines)
    if parciales:
        _escribir_hoja(wb.create_sheet(f"PARCIALES ({len(parciales)})"), parciales,
                       f"PARCIALES — FALTAN ALGUNOS CAJETINES ({len(parciales)})",
                       bg_titulo="7F6000")

    # Hoja NO ENCONTRADOS
    if no_enc:
        _escribir_hoja(wb.create_sheet("SIN CAJETINES"), no_enc,
                       f"SIN NINGÚN CAJETÍN ({len(no_enc)})",
                       bg_titulo="9C0006")

    # Hoja ERRORES
    if errores:
        _escribir_hoja(wb.create_sheet("ERRORES"), errores,
                       f"ERRORES AL ABRIR ({len(errores)})",
                       bg_titulo="C00000")

    wb.save(ruta_salida)
    log.ok(f"Excel guardado: {ruta_salida}")


# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────
def main():
    log.info("=" * 70)
    log.info("  BÚSQUEDA DE CAJETINES OFF/BN — AGP PLANOS TÉCNICOS")
    log.info("=" * 70)
    log.info(f"Ruta base : {RUTA_BASE}")
    log.info(f"Cajetines requeridos ({len(CAJETINES)}):")
    for c in CAJETINES:
        log.info(f"  {c['nombre']}: {' / '.join(c['variantes'])}")
    log.info("-" * 70)

    if not os.path.exists(RUTA_BASE):
        log.error(f"Ruta no accesible: {RUTA_BASE}")
        log.error("Verifica conexión de red y vuelve a intentar.")
        sys.exit(1)

    log.info("\nIMPORTANTE: AutoCAD debe estar abierto (sin archivos) antes de continuar.")
    input("  Presiona Enter cuando AutoCAD esté listo...")

    motor = AutoCAD()

    t0 = time.time()
    resultados = escanear(RUTA_BASE, motor)
    motor.quit()
    duracion = time.time() - t0

    if not resultados:
        log.warn("No se procesó ningún archivo.")
        return

    crear_excel(resultados, ARCHIVO_EXCEL)

    completos = sum(1 for r in resultados if r["estado"] == "COMPLETO")
    parciales = sum(1 for r in resultados if r["estado"] == "PARCIAL")
    no_enc    = sum(1 for r in resultados if r["estado"] == "NO ENCONTRADO")
    errores   = sum(1 for r in resultados if r["estado"] == "ERROR")

    h = int(duracion//3600); m = int((duracion%3600)//60); s = int(duracion%60)

    log.info("\n" + "=" * 70)
    log.info("  RESUMEN")
    log.info("=" * 70)
    log.info(f"  Total escaneados         : {len(resultados)}")
    log.ok  (f"  COMPLETOS (todos cajet.) : {completos}")
    log.warn(f"  PARCIALES (faltan alguno): {parciales}")
    log.info(f"  SIN CAJETINES            : {no_enc}")
    if errores:
        log.error(f"  ERRORES                  : {errores}")
    log.info(f"  Tiempo                   : {h}h {m}m {s}s")
    log.info(f"  Excel                    : {ARCHIVO_EXCEL}")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
