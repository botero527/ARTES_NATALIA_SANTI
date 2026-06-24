#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
AUDITORÍA ARTES AGP
===================
Flujo: ODAFileConverter (DWG→DXF) → ezdxf (lectura) → Azure SQL (update ruta)
Genera reporte Excel con hoja por marca + RESUMEN + ACTUALIZADAS + DUPLICADOS.

Uso:
  python auditoria_artes.py                   # todas las marcas
  python auditoria_artes.py --marca CHEVROLET # solo una marca
  python auditoria_artes.py --reiniciar       # borra checkpoint
  python auditoria_artes.py --solo-excel      # regenera Excel sin procesar DWGs
"""
import os, sys, re, json, time, argparse, shutil, subprocess, tempfile
from datetime import datetime

_DIR_SCRIPT = os.path.dirname(os.path.abspath(__file__))
_DIR_RAIZ   = os.path.dirname(_DIR_SCRIPT)
sys.path.insert(0, _DIR_RAIZ)

# ═══════════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════
RUTA_BASE       = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
CHECKPOINT_FILE = os.path.join(_DIR_SCRIPT, "auditoria_checkpoint.json")
EXCEL_SALIDA    = os.path.join(_DIR_SCRIPT, f"Auditoria_Artes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
CHECKPOINT_CADA = 10
EXCLUIR_CARPETAS = {"OBSOLETO", "OBSOLETOS", "ANTIGUO", "OLD"}

ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

COLOR_K  = 5   # azul ACI
COLOR_K2 = 3   # verde ACI
COLOR_K3 = 1   # rojo ACI

BD_SERVER   = "agpcolombia.database.windows.net"
BD_PORT     = 1433
BD_USER     = "DevIngenieria"
BD_PASSWORD = "HiJE068i0LQVrwA"
BD_DATABASE = "AGP_Ingenieria"

# ═══════════════════════════════════════════════════════════════
#  DEPENDENCIAS
# ═══════════════════════════════════════════════════════════════
try:
    import ezdxf
    from ezdxf.recover import readfile as ezdxf_readfile
except ImportError:
    print("Falta ezdxf.  pip install ezdxf"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Falta openpyxl.  pip install openpyxl"); sys.exit(1)

try:
    import pymssql as _pymssql
except ImportError:
    print("Falta pymssql.  pip install pymssql"); sys.exit(1)

# ═══════════════════════════════════════════════════════════════
#  LOGGER
# ═══════════════════════════════════════════════════════════════
def _ts(): return time.strftime("%H:%M:%S")
def log_info(m):  print(f"{_ts()}  {m}")
def log_warn(m):  print(f"{_ts()}  [!] {m}")
def log_err(m):   print(f"{_ts()}  [X] {m}")
def log_ok(m):    print(f"{_ts()}  [✓] {m}")

# ═══════════════════════════════════════════════════════════════
#  BD
# ═══════════════════════════════════════════════════════════════
class _CursorWrap:
    def __init__(self, c): self._c = c
    def execute(self, sql, p=()):  return self._c.execute(sql.replace("?","%s"), p or ())
    def fetchone(self):            return self._c.fetchone()
    def fetchall(self):            return self._c.fetchall()
    def __getattr__(self, n):      return getattr(self._c, n)

class _ConnWrap:
    def __init__(self, c): self._c = c
    def cursor(self):   return _CursorWrap(self._c.cursor())
    def commit(self):   self._c.commit()
    def rollback(self): self._c.rollback()
    def close(self):    self._c.close()

_BD_CONN = None

def bd_conectar():
    global _BD_CONN
    try:
        if _BD_CONN:
            _BD_CONN.cursor().execute("SELECT 1")
            return _BD_CONN
    except Exception:
        _BD_CONN = None
    _BD_CONN = _ConnWrap(_pymssql.connect(
        server=BD_SERVER, port=BD_PORT, user=BD_USER,
        password=BD_PASSWORD, database=BD_DATABASE,
        timeout=20, login_timeout=20, charset="UTF-8", tds_version="7.3",
    ))
    return _BD_CONN

def bd_validar_y_actualizar(textos, ruta_dwg):
    texto = " ".join(textos)
    vitros  = list(dict.fromkeys(re.findall(r'T-\d{4,6}', texto, re.I)))
    grandes = list(dict.fromkeys(re.findall(r'A-\d{4,6}', texto, re.I)))
    nums_raw = re.findall(r'(?<![TA]-)\b\d{4,6}\b', texto)
    excluir  = {re.sub(r'[TA]-','',v,flags=re.I) for v in vitros+grandes}
    nums = list(dict.fromkeys(n for n in nums_raw if n not in excluir))

    vitro_ok, malla_ok, actualizados, notas, duplicados = [], [], 0, [], []
    try:
        cn = bd_conectar(); cur = cn.cursor()
        for v in vitros:
            cur.execute("SELECT ruta FROM mallas.vitrojet WHERE vitro=?", (v.upper(),))
            row = cur.fetchone()
            if row is not None:
                vitro_ok.append(v.upper())
                ruta_ant = row[0] or ""
                if ruta_ant and ruta_ant.lower() != ruta_dwg.lower():
                    duplicados.append({"codigo": v.upper(), "tipo": "vitro",
                                       "ruta_anterior": ruta_ant, "ruta_nueva": ruta_dwg})
                cur.execute("UPDATE mallas.vitrojet SET ruta=? WHERE vitro=?", (ruta_dwg, v.upper()))
                actualizados += 1
        for m in grandes:
            cur.execute("SELECT ruta_dwg FROM mallas.grandes WHERE codigo=?", (m.upper(),))
            row = cur.fetchone()
            if row is not None:
                malla_ok.append(m.upper())
                ruta_ant = row[0] or ""
                if ruta_ant and ruta_ant.lower() != ruta_dwg.lower():
                    duplicados.append({"codigo": m.upper(), "tipo": "malla_grande",
                                       "ruta_anterior": ruta_ant, "ruta_nueva": ruta_dwg})
                cur.execute("UPDATE mallas.grandes SET ruta_dwg=? WHERE codigo=?", (ruta_dwg, m.upper()))
                actualizados += 1
        for n in nums:
            cur.execute("SELECT ruta_dwg FROM mallas.pequenas WHERE CAST(codigo AS NVARCHAR)=?", (n,))
            row = cur.fetchone()
            if row is not None:
                malla_ok.append(n)
                ruta_ant = row[0] or ""
                if ruta_ant and ruta_ant.lower() != ruta_dwg.lower():
                    duplicados.append({"codigo": n, "tipo": "malla_pequena",
                                       "ruta_anterior": ruta_ant, "ruta_nueva": ruta_dwg})
                cur.execute("UPDATE mallas.pequenas SET ruta_dwg=? WHERE CAST(codigo AS NVARCHAR)=?",
                            (ruta_dwg, n))
                actualizados += 1
        if actualizados:
            cn.commit()
    except Exception as e:
        notas.append(f"BD error: {e}")

    return (
        " / ".join(vitro_ok) or "—",
        " / ".join(malla_ok) or "—",
        actualizados > 0,
        "; ".join(notas),
        duplicados,
    )

# ═══════════════════════════════════════════════════════════════
#  ODA — convertir DWGs a DXF
# ═══════════════════════════════════════════════════════════════
def oda_convertir_carpeta(carpeta_in, carpeta_out, n_archivos=1, timeout=None):
    """
    Convierte todos los DWG/DXF de carpeta_in a DXF en carpeta_out.
    Sintaxis ODA: ODAFileConverter <in> <out> <OutputVersion> <OutputType> <Recursive> <Audit>
    """
    if not os.path.exists(ODA_EXE):
        return False
    os.makedirs(carpeta_out, exist_ok=True)
    # 60s base + 30s por archivo, mínimo 120s
    if timeout is None:
        timeout = max(120, 60 + n_archivos * 30)
    try:
        subprocess.run(
            [ODA_EXE, carpeta_in, carpeta_out, "ACAD2018", "DXF", "0", "1"],
            capture_output=True, timeout=timeout
        )
        return True
    except subprocess.TimeoutExpired:
        log_warn(f"ODA timeout ({timeout}s) en conversión de carpeta")
        return False
    except Exception as e:
        log_warn(f"ODA error: {e}")
        return False

# ═══════════════════════════════════════════════════════════════
#  DETECCIÓN PARABRISAS
# ═══════════════════════════════════════════════════════════════
def es_parabrisas(nombre_archivo):
    stem   = os.path.splitext(nombre_archivo)[0]
    grupos = re.findall(r'\d+', stem)
    if not grupos: return False
    return bool(re.fullmatch(r'0{2,3}', grupos[-1]))

# ═══════════════════════════════════════════════════════════════
#  ANÁLISIS CON EZDXF
# ═══════════════════════════════════════════════════════════════
_RE_FMT = re.compile(r'\{[^}]*\}|\\[A-Za-z][^;]*;|%%.')

def _clean(s):
    return re.sub(r'\s+', ' ', _RE_FMT.sub(' ', s or '')).strip()

def analizar_dxf(ruta_dxf, ruta_dwg_original):
    """Lee un DXF con ezdxf y extrae toda la información."""
    para = es_parabrisas(os.path.basename(ruta_dwg_original))
    r = {
        "ruta": ruta_dwg_original, "archivo": os.path.basename(ruta_dwg_original),
        "es_parabrisas": para,
        "k_ok": False, "k_color": None, "k2_ok": False, "k2_color": None,
        "k3_ok": False, "k3_color": None,
        "logo1_ok": False, "logo2_ok": False,
        "puntos_ok": False, "puntos_estado": "—",
        "vitro": "—", "malla": "—", "bd_actualizado": False, "duplicados": [],
        "estado": "OK", "notas": [], "error": None,
    }
    try:
        doc, _ = ezdxf_readfile(ruta_dxf)
        msp    = doc.modelspace()

        # Layers → color ACI
        layers = {}
        for lyr in doc.layers:
            layers[lyr.dxf.name.upper().strip()] = abs(lyr.color)

        textos          = []
        layers_con_ents = set()
        hatch_puntos    = False
        trazo_puntos    = False
        _bloques_vistos = set()

        def _texto_ent(e):
            t = e.dxftype()
            lyr = e.dxf.layer.upper().strip() if e.dxf.hasattr("layer") else ""
            if lyr:
                layers_con_ents.add(lyr)
            if t in ("TEXT", "ATTRIB", "ATTDEF"):
                try:
                    s = _clean(e.dxf.text)
                    if s: textos.append(s)
                except Exception: pass
            elif t == "MTEXT":
                try:
                    s = _clean(e.plain_mtext())
                    if s: textos.append(s)
                except Exception: pass
            elif t == "TABLE":
                try:
                    for row in range(e.dxf.rows):
                        for col in range(e.dxf.columns):
                            try:
                                s = _clean(e.get_cell_value(row, col))
                                if s: textos.append(s)
                            except Exception: pass
                except Exception: pass
            if "PUNTOS" in lyr:
                nonlocal hatch_puntos, trazo_puntos
                if t == "HATCH": hatch_puntos = True
                elif t in ("LWPOLYLINE","POLYLINE","LINE","SPLINE"): trazo_puntos = True

        def _leer_blk(nombre):
            if nombre in _bloques_vistos: return
            _bloques_vistos.add(nombre)
            try:
                for e in doc.blocks[nombre]:
                    _texto_ent(e)
                    if e.dxftype() == "INSERT":
                        try:
                            for a in e.attribs: _texto_ent(a)
                        except Exception: pass
                        try: _leer_blk(e.dxf.name)
                        except Exception: pass
            except Exception: pass

        for e in msp:
            _texto_ent(e)
            if e.dxftype() == "INSERT":
                try:
                    for a in e.attribs: _texto_ent(a)
                except Exception: pass
                try: _leer_blk(e.dxf.name)
                except Exception: pass

        # Validar K / K2 / K3
        for key, color_esp, campo_ok, campo_color in [
            ("K",  COLOR_K,  "k_ok",  "k_color"),
            ("K2", COLOR_K2, "k2_ok", "k2_color"),
            ("K3", COLOR_K3, "k3_ok", "k3_color"),
        ]:
            color_real = layers.get(key)
            r[campo_color] = color_real
            if key in layers_con_ents and color_real == color_esp:
                r[campo_ok] = True
            elif key in layers_con_ents and color_real != color_esp:
                r["notas"].append(f"Layer {key} color incorrecto ({color_real}≠{color_esp})")
            elif key in layers and key not in layers_con_ents:
                r["notas"].append(f"Layer {key} vacío")
            else:
                r["notas"].append(f"Falta layer {key}")

        # Logo (aplica a todos) — cualquier layer que contenga "LOGO" (case-insensitive)
        layers_upper = {l.upper() for l in layers_con_ents}
        logos = [l for l in layers_upper if "LOGO" in l]
        logos2 = [l for l in logos if "2" in l]
        logo_ok = len(logos) > 0
        r["logo1_ok"] = logo_ok
        r["logo2_ok"] = len(logos2) > 0
        if not logo_ok:
            r["notas"].append("Falta layer LOGO")

        # Puntos (solo parabrisas)
        if para:
            tiene_puntos = any("PUNTOS" in l for l in layers_con_ents)
            r["puntos_ok"] = tiene_puntos
            if not tiene_puntos:
                r["puntos_estado"] = "FALTA"
                r["notas"].append("Parabrisas sin layer PUNTOS")
            elif hatch_puntos and not trazo_puntos:
                r["puntos_estado"] = "OK relleno"
            elif trazo_puntos and not hatch_puntos:
                r["puntos_estado"] = "DESACTUALIZADO (trazo suelto)"
                r["notas"].append("PUNTOS: trazo suelto sin relleno")
            else:
                r["puntos_estado"] = "MIXTO"

        # BD
        vitro, malla, bd_ok, notas_bd, dups = bd_validar_y_actualizar(textos, ruta_dwg_original)
        r["vitro"] = vitro; r["malla"] = malla
        r["bd_actualizado"] = bd_ok; r["duplicados"] = dups
        if notas_bd: r["notas"].append(notas_bd)
        if dups: r["notas"].append(f"DUPLICADO: {', '.join(d['codigo'] for d in dups)}")

        # Estado
        tiene_algo = any(l in layers_con_ents for l in ("K","K2","K3","LOGO","LOGO1","LOGO2"))
        if not tiene_algo and vitro == "—" and malla == "—":
            r["estado"] = "SIN DATOS"
            r["notas"]  = ["Sin layers/texto estándar"]
        else:
            errores = [n for n in r["notas"] if any(p in n for p in ("Falta","incorrecto","DESACTUALIZADO"))]
            r["estado"] = "INCOMPLETO" if errores else "OK"

    except Exception as e:
        r["estado"] = "ERROR"
        r["error"]  = str(e)
        r["notas"]  = [f"Error ezdxf: {e}"]

    r["notas"] = " | ".join(r["notas"]) if r["notas"] else ""
    return r

# ═══════════════════════════════════════════════════════════════
#  RECORRER ÁRBOL
# ═══════════════════════════════════════════════════════════════
def _excluir(nombre):
    return any(ex in nombre.upper() for ex in EXCLUIR_CARPETAS)

def _extraer_contexto(ruta_artes, ruta_marca):
    partes = []
    p = os.path.dirname(ruta_artes)
    for _ in range(6):
        cab, cola = os.path.split(p)
        if not cola or p == ruta_marca or p == cab: break
        partes.insert(0, cola); p = cab
    if len(partes) >= 2: return partes[-2], partes[-1]
    elif len(partes) == 1: return partes[0], ""
    return "", ""

def listar_carpetas_artes(ruta_base, solo_marca=None):
    """Genera (marca, vehiculo, version, carpeta_artes, [archivos_P])."""
    try:
        marcas = sorted(os.listdir(ruta_base))
    except Exception as e:
        log_err(f"No se puede leer {ruta_base}: {e}"); return

    for marca in marcas:
        if solo_marca and marca.upper() != solo_marca.upper(): continue
        ruta_marca = os.path.join(ruta_base, marca)
        if not os.path.isdir(ruta_marca): continue

        for dirpath, dirnames, filenames in os.walk(ruta_marca):
            dirnames[:] = sorted(d for d in dirnames if not _excluir(d))
            if os.path.basename(dirpath).upper() != "ARTES": continue

            vehiculo, version = _extraer_contexto(dirpath, ruta_marca)

            # Archivos P* directamente en ARTES/
            p_artes = [os.path.join(dirpath, f) for f in sorted(filenames)
                       if f.lower().endswith((".dwg",".dxf")) and f.upper().startswith("P") and not _excluir(f)]

            # Subcarpeta BN
            p_bn = []
            for sub in sorted(os.listdir(dirpath)):
                if sub.upper() == "BN":
                    carpeta_bn = os.path.join(dirpath, sub)
                    if os.path.isdir(carpeta_bn):
                        p_bn = [os.path.join(carpeta_bn, f)
                                for f in sorted(os.listdir(carpeta_bn))
                                if f.lower().endswith((".dwg",".dxf")) and f.upper().startswith("P") and not _excluir(f)]

            for grupo in [p_artes, p_bn]:
                if grupo:
                    yield marca, vehiculo, version, os.path.dirname(grupo[0]), grupo

# ═══════════════════════════════════════════════════════════════
#  CHECKPOINT
# ═══════════════════════════════════════════════════════════════
def cp_cargar():
    if not os.path.exists(CHECKPOINT_FILE): return {}
    try:
        with open(CHECKPOINT_FILE, encoding="utf-8") as f: return json.load(f)
    except Exception: return {}

def cp_guardar(data):
    try:
        with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log_warn(f"No se pudo guardar checkpoint: {e}")

# ═══════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════
COLS = [
    ("Archivo",28),("Marca",14),("Vehículo",36),("Versión",10),
    ("Parabrisas",12),("K",8),("K2",8),("K3",8),("Logo",8),("Logo2",8),
    ("Puntos",22),("Vitro",14),("Malla",20),("BD Update",10),
    ("Estado",14),("Notas",55),("Ruta",70),
]
C_HDR="1F4E79"; C_OK="C6EFCE"; C_ERR="FFCCCC"; C_WARN="FFE699"
C_PARA="D6E4F0"; C_ALT="F2F2F2"; C_SD="E2E2E2"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def _borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _escribir_filas(ws, filas, es_actualizadas=False):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for ci, (titulo, ancho) in enumerate(COLS, 1):
        cel = ws.cell(row=1, column=ci, value=titulo)
        cel.font = _font(bold=True, color="FFFFFF", size=10)
        cel.fill = _fill(C_HDR)
        cel.alignment = _center()
        cel.border = _borde()
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    for ri, f in enumerate(filas, 2):
        estado = f.get("estado", "")
        para   = f.get("es_parabrisas", False)
        if es_actualizadas:
            fila_color = "E2EFDA" if ri % 2 == 0 else "F0FAF0"
        else:
            fila_color = (C_ERR  if estado == "ERROR"
                          else C_WARN if estado == "INCOMPLETO"
                          else C_SD   if estado == "SIN DATOS"
                          else C_PARA if para
                          else (C_ALT  if ri % 2 == 0 else "FFFFFF"))
        vals = [
            f.get("archivo",""), f.get("marca",""), f.get("vehiculo",""), f.get("version",""),
            "✔" if para else "",
            "✔" if f.get("k_ok")  else "✘",
            "✔" if f.get("k2_ok") else "✘",
            "✔" if f.get("k3_ok") else "✘",
            "✔" if f.get("logo1_ok") else "✘",
            "✔" if f.get("logo2_ok") else "—",
            f.get("puntos_estado","—"),
            f.get("vitro","—"), f.get("malla","—"),
            "✔" if f.get("bd_actualizado") else "—",
            estado, f.get("notas",""), f.get("ruta",""),
        ]
        for ci, val in enumerate(vals, 1):
            cel = ws.cell(row=ri, column=ci, value=val)
            cel.fill = _fill(fila_color); cel.border = _borde(); cel.alignment = _left()
            if ci in (6,7,8,9,10,14):
                cel.alignment = _center()
                if str(val) == "✘": cel.font = _font(bold=True, color="C00000")
                elif str(val) == "✔": cel.font = _font(bold=True, color="375623")

def generar_excel(filas_por_marca, ruta_excel):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    totales_resumen = []
    todas_filas = []

    for marca, filas in sorted(filas_por_marca.items()):
        ws = wb.create_sheet(title=marca[:28])
        _escribir_filas(ws, filas)
        todas_filas.extend(filas)
        ok  = sum(1 for f in filas if f.get("estado") == "OK")
        inc = sum(1 for f in filas if f.get("estado") == "INCOMPLETO")
        err = sum(1 for f in filas if f.get("estado") not in ("OK","INCOMPLETO","SIN DATOS"))
        sd  = sum(1 for f in filas if f.get("estado") == "SIN DATOS")
        bd  = sum(1 for f in filas if f.get("bd_actualizado"))
        total = len(filas)
        totales_resumen.append({"marca": marca, "total": total, "ok": ok,
                                 "incompleto": inc, "error": err, "sin_datos": sd,
                                 "bd_upd": bd,
                                 "pct": f"{ok/total*100:.1f}%" if total else "0%"})

    # RESUMEN
    ws = wb.create_sheet(title="RESUMEN", index=0)
    ws.freeze_panes = "A2"
    hdr = ["Marca","Total","OK","Incompleto","Sin datos","Error","BD Update","% OK"]
    anc = [22,9,9,12,12,9,12,9]
    for ci,(h,w) in enumerate(zip(hdr,anc),1):
        cel = ws.cell(row=1, column=ci, value=h)
        cel.font = _font(bold=True, color="FFFFFF", size=10)
        cel.fill = _fill(C_HDR); cel.alignment = _center(); cel.border = _borde()
        ws.column_dimensions[get_column_letter(ci)].width = w
    tt=tok=tinc=terr=tsd=tbd=0
    for ri, t in enumerate(totales_resumen, 2):
        for ci, val in enumerate([t["marca"],t["total"],t["ok"],t["incompleto"],
                                   t["sin_datos"],t["error"],t["bd_upd"],t["pct"]], 1):
            cel = ws.cell(row=ri, column=ci, value=val)
            cel.border = _borde(); cel.alignment = _center() if ci>1 else _left()
        tt+=t["total"]; tok+=t["ok"]; tinc+=t["incompleto"]
        terr+=t["error"]; tsd+=t["sin_datos"]; tbd+=t["bd_upd"]
    ri_t = len(totales_resumen)+2
    for ci, val in enumerate(["TOTAL",tt,tok,tinc,tsd,terr,tbd,
                               f"{tok/tt*100:.1f}%" if tt else "0%"],1):
        cel = ws.cell(row=ri_t, column=ci, value=val)
        cel.font = _font(bold=True); cel.fill = _fill("D9E1F2")
        cel.border = _borde(); cel.alignment = _center() if ci>1 else _left()

    # ACTUALIZADAS
    actualizadas = [f for f in todas_filas if f.get("bd_actualizado")]
    if actualizadas:
        ws_act = wb.create_sheet(title="ACTUALIZADAS")
        _escribir_filas(ws_act, actualizadas, es_actualizadas=True)
        log_ok(f"Hoja ACTUALIZADAS: {len(actualizadas)} artes")

    # DUPLICADOS
    todos_dups = []
    for f in todas_filas:
        for d in f.get("duplicados", []):
            todos_dups.append({**d, "marca": f.get("marca",""), "vehiculo": f.get("vehiculo",""),
                               "version": f.get("version",""), "archivo": f.get("archivo","")})
    if todos_dups:
        ws_dup = wb.create_sheet(title="DUPLICADOS")
        ws_dup.freeze_panes = "A2"
        hdr_d = ["Código","Tipo","Marca","Vehículo","Versión","Archivo","Ruta anterior","Ruta nueva"]
        anc_d = [14,14,16,36,10,36,70,70]
        for ci,(h,w) in enumerate(zip(hdr_d,anc_d),1):
            cel = ws_dup.cell(row=1, column=ci, value=h)
            cel.font = _font(bold=True, color="FFFFFF", size=10)
            cel.fill = _fill("7B2D00"); cel.alignment = _center(); cel.border = _borde()
            ws_dup.column_dimensions[get_column_letter(ci)].width = w
        for ri, d in enumerate(todos_dups, 2):
            for ci, val in enumerate([d["codigo"],d["tipo"],d["marca"],d["vehiculo"],
                                       d["version"],d["archivo"],d["ruta_anterior"],d["ruta_nueva"]], 1):
                cel = ws_dup.cell(row=ri, column=ci, value=val)
                cel.fill = _fill("FFF2CC" if ri%2==0 else "FFFAF0")
                cel.border = _borde(); cel.alignment = _left()

    wb.save(ruta_excel)
    log_ok(f"Excel guardado: {ruta_excel}")

# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Auditoría artes AGP")
    parser.add_argument("--marca",      default=None)
    parser.add_argument("--reiniciar",  action="store_true")
    parser.add_argument("--solo-excel", action="store_true")
    args = parser.parse_args()

    if args.reiniciar and os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        log_info("Checkpoint borrado.")

    cp = cp_cargar()

    if args.solo_excel:
        filas_por_marca = {}
        for key, filas in cp.items():
            marca = key.split("::")[0] if "::" in key else key
            filas_por_marca.setdefault(marca, []).extend(filas)
        generar_excel(filas_por_marca, EXCEL_SALIDA)
        return

    if not os.path.exists(ODA_EXE):
        log_err(f"ODA File Converter no encontrado: {ODA_EXE}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  AUDITORÍA ARTES AGP — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Base: {RUTA_BASE}")
    if args.marca: print(f"  Marca: {args.marca}")
    print(f"{'='*60}\n")

    filas_por_marca = {}
    for key, filas in cp.items():
        marca = key.split("::")[0]
        filas_por_marca.setdefault(marca, []).extend(filas)

    buffer_carpeta = []
    clave_carpeta  = None
    n_archivo      = 0
    n_desde_cp     = 0

    try:
        for marca, vehiculo, version, carpeta, archivos in listar_carpetas_artes(RUTA_BASE, args.marca):
            clave = f"{marca}::{vehiculo}::{version}::{carpeta}"

            # Filtrar ya procesados
            ya_procesados = {f["ruta"] for f in cp.get(clave, [])}
            pendientes = [a for a in archivos if a not in ya_procesados]
            if not pendientes:
                continue

            # Cambio de grupo → guardar buffer
            if clave != clave_carpeta:
                if buffer_carpeta and clave_carpeta:
                    cp.setdefault(clave_carpeta, []).extend(buffer_carpeta)
                    cp_guardar(cp)
                buffer_carpeta = []
                clave_carpeta  = clave
                log_info(f"→ {marca} / {vehiculo} / {version}")

            # Convertir carpeta entera con ODA a DXF en carpeta temp
            tmp_in  = tempfile.mkdtemp(prefix="oda_in_")
            tmp_out = tempfile.mkdtemp(prefix="oda_out_")
            try:
                # Copiar solo los pendientes al tmp_in
                for ruta_orig in pendientes:
                    try:
                        shutil.copy2(ruta_orig, os.path.join(tmp_in, os.path.basename(ruta_orig)))
                    except Exception as ce:
                        log_warn(f"  No se pudo copiar {os.path.basename(ruta_orig)}: {ce}")

                log_info(f"  Convirtiendo {len(pendientes)} archivo(s) con ODA...")
                oda_ok = oda_convertir_carpeta(tmp_in, tmp_out, n_archivos=len(pendientes))
                if not oda_ok:
                    log_warn("  ODA falló — archivos quedarán como ERROR")

                # Procesar cada archivo
                for ruta_orig in pendientes:
                    nombre = os.path.basename(ruta_orig)
                    stem   = os.path.splitext(nombre)[0]
                    ruta_dxf = os.path.join(tmp_out, stem + ".dxf")

                    n_archivo  += 1
                    n_desde_cp += 1
                    log_info(f"  [{n_archivo}] {nombre}")

                    if os.path.exists(ruta_dxf):
                        resultado = analizar_dxf(ruta_dxf, ruta_orig)
                    else:
                        # ODA no pudo convertir este archivo
                        resultado = {
                            "ruta": ruta_orig, "archivo": nombre,
                            "es_parabrisas": es_parabrisas(nombre),
                            "k_ok": False, "k_color": None, "k2_ok": False, "k2_color": None,
                            "k3_ok": False, "k3_color": None, "logo1_ok": False, "logo2_ok": False,
                            "puntos_ok": False, "puntos_estado": "—",
                            "vitro": "—", "malla": "—", "bd_actualizado": False, "duplicados": [],
                            "estado": "ERROR", "notas": "ODA no pudo convertir", "error": "oda_fail",
                        }

                    resultado["marca"]    = marca
                    resultado["vehiculo"] = vehiculo
                    resultado["version"]  = version

                    buffer_carpeta.append(resultado)
                    filas_por_marca.setdefault(marca, []).append(resultado)

                    # Log resultado
                    est  = resultado["estado"]
                    r    = resultado
                    def _c(ok): return "✔" if ok else "✘"
                    lay = (f"K={_c(r['k_ok'])}({r['k_color']})  "
                           f"K2={_c(r['k2_ok'])}({r['k2_color']})  "
                           f"K3={_c(r['k3_ok'])}({r['k3_color']})  "
                           f"Logo={_c(r['logo1_ok'])}")
                    if r["es_parabrisas"]: lay += f"  Puntos={r['puntos_estado']}"
                    bd = ""
                    if r["vitro"] != "—" or r["malla"] != "—":
                        bd = (f"  vitro={r['vitro']}  malla={r['malla']}  "
                              f"BD={'✔' if r['bd_actualizado'] else '✘'}")
                    else:
                        bd = "  vitro=—  malla=—"
                    prefijo = ("  [OK]" if est=="OK" else
                               "  [~~]" if est=="SIN DATOS" else
                               "  [!!]" if est=="INCOMPLETO" else "  [XX]")
                    notas_str = f"\n      ⚠  {r['notas']}" if r["notas"] else ""
                    log_info(f"{prefijo}  {lay}{bd}{notas_str}")

                    if n_desde_cp >= CHECKPOINT_CADA:
                        cp.setdefault(clave_carpeta, []).extend(buffer_carpeta)
                        buffer_carpeta = []
                        cp_guardar(cp)
                        log_info(f"  [checkpoint — {n_archivo} archivos]")
                        n_desde_cp = 0

            finally:
                shutil.rmtree(tmp_in,  ignore_errors=True)
                shutil.rmtree(tmp_out, ignore_errors=True)

    except KeyboardInterrupt:
        log_warn("Interrumpido.")
    finally:
        if buffer_carpeta and clave_carpeta:
            cp.setdefault(clave_carpeta, []).extend(buffer_carpeta)
        cp_guardar(cp)

    print(f"\n{'='*60}")
    log_ok(f"Procesados: {n_archivo} archivos")
    log_info("Generando Excel...")
    generar_excel(filas_por_marca, EXCEL_SALIDA)
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
